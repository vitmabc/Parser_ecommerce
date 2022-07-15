from configparser import ConfigParser
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import datetime
import sqlite3
import numpy as np
from sqlalchemy import create_engine
import pandas as pds


def main():
    def config(filename='database.ini', section='alchemyEngine'):  # Подключение к Postgres
        # create a parser
        parser = ConfigParser()
        # read config file
        parser.read(filename)
        db = {}
        if parser.has_section(section):
            parameters = parser.items(section)
            for param in parameters:
                db[param[0]] = param[1]
        else:
            raise Exception('Section {0} not found in the {1} file'.format(section, filename))

        return db


    def modify_piramida24_table(name):  # Поиск артикулов в фрейме Пирамиды и запись их в отдельную колонку sku
        for row in dataFrame_article:
            mask = dataFrame_piramida[dataFrame_piramida[name].str.contains(row, regex=False)]
            index = mask.index.tolist()
            for b in index:
                if b:
                    if dataFrame_piramida.loc[b, 'sku_piramida'] == '':
                        dataFrame_piramida.loc[[b], 'sku_piramida'] = row
                    elif dataFrame_piramida.loc[b, 'sku_piramida'] == row:
                        pass
                    else:
                        dataFrame_piramida.loc[[b], 'sku_piramida'] = row + ', ' + dataFrame_piramida.loc[
                            [b], 'sku_piramida']
                else:
                    pass


    def modify_gazkomplekt_table(name):  # Поиск артикулов в фрейме gazkomplekt и запись их в отдельную колонку sku
        for row in dataFrame_article:
            mask = dataFrame_gazkomplekt[dataFrame_gazkomplekt[name].str.contains(row, regex=False)]
            index = mask.index.tolist()
            for b in index:
                if b:
                    if dataFrame_gazkomplekt.loc[b, 'sku_gazkomplekt'] == '':
                        dataFrame_gazkomplekt.loc[[b], 'sku_gazkomplekt'] = row
                    elif dataFrame_gazkomplekt.loc[b, 'sku_gazkomplekt'] == row:
                        pass
                    else:
                        dataFrame_gazkomplekt.loc[[b], 'sku_gazkomplekt'] = row + ', ' + dataFrame_gazkomplekt.loc[
                            [b], 'sku_gazkomplekt']
                else:
                    pass

    def formats(file, name):  # Форматирование Excel файла
        wb = load_workbook(filename=file)
        ws = wb[name]
        ws['A1'] = str('Update date: ' + datetime.datetime.now().strftime("%d-%m-%Y %H:%M"))
        ws['A1'].alignment = Alignment(vertical='center')
        ws['A1'].font = Font(size=7)
        ws.column_dimensions['A'].width = 39
        ws.column_dimensions['B:J'].width = 13
        # ws.column_dimensions['C'].width = 13
        # ws.column_dimensions['D'].width = 13
        # ws.column_dimensions['E'].width = 13
        # ws.column_dimensions['F'].width = 13
        # выравнивание текста
        for row in ws[3:ws.max_row]:  # skip the header
            cell = row[1]  # column H
            cell.alignment = Alignment(horizontal='center')
            # cell = row[2]  # column H
            # cell.alignment = Alignment(horizontal='center')
            cell = row[3]
            cell.number_format = "0.00"
            cell = row[5]
            cell.number_format = "0.00"
            cell = row[7]
            cell.number_format = "0.00"
            cell = row[9]
            cell.number_format = "0.00"
        wb.save(filename=file)


    dbConnection = sqlite3.connect('database.db', check_same_thread=True)

    dataFrame_teploz = pds.read_sql(
        "select description as description_teploz,url as url_teploz, id as id_teploz,sku as sku_teploz,\
        article as article_teploz,name as name_teploz,price as price_teploz from "
        "\"teplozapchast_catalog\"",
        dbConnection)
    dataFrame_kotel = pds.read_sql(
        "select description as description_kotel_kr,url as url_kotel, id as id_kotel_kr,sku as sku_kotel_kr,\
        article as article_kotel_kr,name as name_kotel_kr,price as price_kotel_kr from "
        "\"kotel_kr_catalog\"",
        dbConnection)
    dataFrame_piramida = pds.read_sql(
        "select description as description_piramida,url as url_piramida, name,id as id_piramida,sku as sku_piramida,\
        article as article_piramida,name as name_piramida,"
        "price as price_piramida from \"piramida_catalog\"",
        dbConnection)

    dataFrame_gazkomplekt = pds.read_sql(
        "select description as description_gazc,url as url_gazkomplekt, name,id as id_gazkomplekt,sku as sku_gazkomplekt,\
        article as article_gazkomplekt,"
        "name as name_gazkomplekt, "
        "price as price_gazkomplekt from \"gazkomplekt_catalog\"",
        dbConnection)

    # Разделение товаров на отдельные строки с отдельным артикулом, если у товара несколько артикулов и они перечислены
    # через запятую
    dataFrame_teploz = dataFrame_teploz["sku_teploz"].str.split(",").explode().str.strip().to_frame().combine_first(
        dataFrame_teploz)
    dataFrame_article = pds.concat([dataFrame_kotel['sku_kotel_kr'], dataFrame_teploz['sku_teploz']],
                                   ignore_index=False).drop_duplicates()

    list_error_name = ['G1/2', 'G3/4', '- 100', '- 250', 'G3/8', 'G1/2']
    dataFrame_article = dataFrame_article.loc[~dataFrame_article.isin(list_error_name)]

    # Поиск артикулов в фрейме Пирамиды по полю Описание и Название и запись их в отдельную колонку sku
    modify_piramida24_table('description_piramida')
    modify_piramida24_table('name')
    modify_gazkomplekt_table('name')


    dataFrame_piramida = dataFrame_piramida["sku_piramida"].str.split(",").explode().str.strip().to_frame().combine_first(
        dataFrame_piramida)
    dataFrame_gazkomplekt = dataFrame_gazkomplekt["sku_gazkomplekt"].str.split(
        ",").explode().str.strip().to_frame().combine_first(
        dataFrame_gazkomplekt)

    dataFrame_piramida['SKU_p'] = dataFrame_piramida.apply(lambda row: row['sku_piramida'].lstrip() if row['sku_piramida'] !='' else row['article_piramida'].lstrip(), axis=1)
    dataFrame_gazkomplekt['SKU_g'] = dataFrame_gazkomplekt.apply(lambda row: row['sku_gazkomplekt'] if row['sku_gazkomplekt'] !='' else '', axis=1)


    df = dataFrame_kotel.merge(dataFrame_teploz, how='outer', left_on='sku_kotel_kr', right_on='sku_teploz')
    df['Name_temp1'] = df.apply(lambda row: row['name_kotel_kr'].lstrip() if row['name_kotel_kr']==row['name_kotel_kr'] else row['name_teploz'].lstrip(), axis=1)
    df['SKU_temp1'] = df.apply(lambda row: row['sku_kotel_kr'] if row['sku_kotel_kr']==row['sku_kotel_kr'] else row['sku_teploz'], axis=1)


    df = df.merge(dataFrame_piramida, how='outer', left_on=['SKU_temp1'], right_on='SKU_p')
    df['Name_temp2'] = df.apply(lambda row: row['Name_temp1'].lstrip() if row['Name_temp1']==row['Name_temp1'] else row['name'].lstrip(), axis=1)
    df['SKU_temp2'] = df.apply(lambda row: row['SKU_temp1'] if row['SKU_temp1']==row['SKU_temp1'] else row['SKU_p'], axis=1)


    df = df.merge(dataFrame_gazkomplekt, how='outer', left_on=['SKU_temp2'], right_on='SKU_g')
    df['Name'] = df.apply(lambda row: row['Name_temp2'].lstrip() if row['Name_temp2']==row['Name_temp2'] else row['name_gazkomplekt'].lstrip(), axis=1)
    df['SKU'] = df.apply(lambda row: row['SKU_temp2'] if row['SKU_temp2']==row['SKU_temp2'] else row['SKU_g'], axis=1)

    df = df.drop_duplicates()


    DF = df.reindex(columns=['Name', 'SKU', 'article_kotel_kr', 'price_kotel_kr', 'article_teploz', 'price_teploz', 'article_piramida', 'price_piramida',
                        'sku_gazkomplekt', 'price_gazkomplekt'])
    DF1 = DF.sort_values(by='Name')
    with pds.ExcelWriter('result.xlsx', engine="openpyxl") as wb:
        DF1.to_excel(wb, sheet_name='name', startrow=1, index=False)

    with pds.ExcelWriter('result2.xlsx', engine="openpyxl") as wb:
        dataFrame_article.to_excel(wb, sheet_name='name', startrow=1, index=False)

    with pds.ExcelWriter('resultGaz.xlsx', engine="openpyxl") as wb:
        dataFrame_gazkomplekt.to_excel(wb, sheet_name='name', startrow=1, index=False)

    formats('result.xlsx','name')

if __name__ =='__main__':
    main()