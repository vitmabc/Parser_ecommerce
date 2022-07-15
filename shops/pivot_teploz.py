from configparser import ConfigParser
import pandas as pds
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import numpy as np
import sqlite3
import datetime


def start():
    def config(filename='database.ini', section='alchemyEngine'):  # Подключение к Postgres
        # create a parser
        parser = ConfigParser()
        # read config file
        parser.read(filename)
        db = {}
        if parser.has_section(section):
            parameters = parser.items(section)
            for param in parameters:
                db[param[0]] = param[1]
        else:
            raise Exception('Section {0} not found in the {1} file'.format(section, filename))

        return db

    def modify_piramida24_table(name):  # Поиск артикулов в фрейме Пирамиды и запись их в отдельную колонку sku
        for row in dataFrame_article:
            mask = dataFrame_piramida[dataFrame_piramida[name].str.contains(row, regex=False)]
            index = mask.index.tolist()
            for b in index:
                if b:
                    if dataFrame_piramida.loc[b, 'sku_piramida'] == '':
                        dataFrame_piramida.loc[[b], 'sku_piramida'] = row
                    elif dataFrame_piramida.loc[b, 'sku_piramida'] == row:
                        pass
                    else:
                        dataFrame_piramida.loc[[b], 'sku_piramida'] = row + ', ' + dataFrame_piramida.loc[
                            [b], 'sku_piramida']
                else:
                    pass

    def modify_gazkomplekt_table(name):  # Поиск артикулов в фрейме gazkomplekt и запись их в отдельную колонку sku
        for row in dataFrame_article:
            mask = dataFrame_gazkomplekt[dataFrame_gazkomplekt[name].str.contains(row, regex=False)]
            index = mask.index.tolist()
            for b in index:
                if b:
                    if dataFrame_gazkomplekt.loc[b, 'sku_gazkomplekt'] == '':
                        dataFrame_gazkomplekt.loc[[b], 'sku_gazkomplekt'] = row
                    elif dataFrame_gazkomplekt.loc[b, 'sku_gazkomplekt'] == row:
                        pass
                    else:
                        dataFrame_gazkomplekt.loc[[b], 'sku_gazkomplekt'] = row + ', ' + dataFrame_gazkomplekt.loc[
                            [b], 'sku_gazkomplekt']
                else:
                    pass

    def merge_table(dataframe1, dataframe2, left, right):  # Объединение двух таблиц INNER
        tempdataframe1 = dataframe1
        tempdataframe1[left].replace('', np.nan, inplace=True)
        tempdataframe1.dropna(subset=[left], inplace=True)
        merged_table = tempdataframe1.merge(dataframe2, left_on=left, right_on=right)

        return merged_table

    def drop_duplicate(res, id1, id2):  # Удаление дубликатов в объединенной таблице
        res = res.drop_duplicates([id1, id2])
        return res

    def write_excel(res, filename, name, **kwargs):  # Запись в Excel файл
        with pds.ExcelWriter(filename, engine="openpyxl", **kwargs) as wb:
            res.to_excel(wb, sheet_name=name, startrow=1, index=False)

    def formats(file, name):  # Форматирование Excel файла
        wb = load_workbook(filename=file)
        ws = wb[name]
        ws['A1'] = str('Update date: ' + datetime.datetime.now().strftime("%d-%m-%Y %H:%M"))
        ws['A1'].alignment = Alignment(vertical='center')
        ws['A1'].font = Font(size=7)
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['B'].width = 51
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 51
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        # выравнивание текста
        for row in ws[3:ws.max_row]:  # skip the header
            cell = row[0]  # column H
            cell.alignment = Alignment(horizontal='center')
            cell = row[2]  # column H
            cell.alignment = Alignment(horizontal='center')
            cell = row[4]
            cell.number_format = "0.00"
            cell = row[5]
            cell.number_format = "0.00"
        wb.save(filename=file)

    def kotel_kr_teplozap():
        # Сравнение kotel_kr и teplozapchast
        res_t = merge_table(dataFrame_kotel, dataFrame_teploz, 'sku_kotel_kr', "sku_teploz")
        res_t = drop_duplicate(res_t, 'id_kotel_kr', 'id_teploz')
        res_t = res_t.reindex(
            columns=['sku_kotel_kr', 'name_kotel_kr', 'sku_teploz', 'name_teploz', 'price_kotel_kr', 'price_teploz'])

        kotel_kr_lower = res_t[res_t['price_kotel_kr'] < res_t['price_teploz']]
        teplozap_lower = res_t[res_t['price_kotel_kr'] > res_t['price_teploz']]
        filename = 'cross_kotel_kr_teplozapchast.xlsx'
        write_excel(res_t, filename, 'cross_table_kotel-teploz')
        write_excel(kotel_kr_lower, filename, 'kotel_kr_lower', mode="a", if_sheet_exists='replace')
        write_excel(teplozap_lower, filename, 'teplozap_lower', mode="a", if_sheet_exists='replace')
        formats(filename, 'cross_table_kotel-teploz')
        formats(filename, 'kotel_kr_lower')
        formats(filename, 'teplozap_lower')

    def teplozap_piramida():
        # Сравнение teplozapchast и piramida
        res_t_p = merge_table(dataFrame_teploz, dataFrame_piramida, 'sku_teploz', "sku_piramida")
        res_t_p1 = merge_table(dataFrame_teploz, dataFrame_piramida, 'sku_teploz', "article_piramida")
        res_t_p = pds.concat([res_t_p, res_t_p1])
        res_t_p = drop_duplicate(res_t_p, 'id_teploz', 'id_piramida')
        res_t_p = res_t_p.reindex(
            columns=['sku_teploz', 'name_teploz', 'article_piramida', 'name_piramida', 'price_teploz',
                     'price_piramida'])
        teplozap_lower = res_t_p[res_t_p['price_teploz'] < res_t_p['price_piramida']]
        piramida_lower = res_t_p[res_t_p['price_teploz'] > res_t_p['price_piramida']]
        filename = 'cross_teplozap_piramida.xlsx'
        write_excel(res_t_p, filename, 'cross_table_teplozap-piramida')
        write_excel(teplozap_lower, filename, 'teplozap_lower', mode="a", if_sheet_exists='replace')
        write_excel(piramida_lower, filename, 'piramida_lower', mode="a", if_sheet_exists='replace')
        formats(filename, 'cross_table_teplozap-piramida')
        formats(filename, 'teplozap_lower')
        formats(filename, 'piramida_lower')

    def teplozap_gazkompl():
        # Сравнение kotel_kr и teplozapchast
        res_t = merge_table(dataFrame_gazkomplekt, dataFrame_teploz, 'sku_gazkomplekt', "sku_teploz")
        res_t = drop_duplicate(res_t, 'id_gazkomplekt', 'id_teploz')
        res_t = res_t.reindex(
            columns=['sku_gazkomplekt', 'name_gazkomplekt', 'sku_teploz', 'name_teploz', 'price_gazkomplekt',
                     'price_teploz'])

        gazk_lower = res_t[res_t['price_gazkomplekt'] < res_t['price_teploz']]
        teplozap_lower = res_t[res_t['price_gazkomplekt'] > res_t['price_teploz']]
        filename = 'cross_gazkomplekt_teplozapchast.xlsx'
        write_excel(res_t, filename, 'cross_gazkomplekt_teplozapchast')
        write_excel(gazk_lower, filename, 'gazkompl_lower', mode="a", if_sheet_exists='replace')
        write_excel(teplozap_lower, filename, 'teplozap_lower', mode="a", if_sheet_exists='replace')
        formats(filename, 'cross_gazkomplekt_teplozapchast')
        formats(filename, 'gazkompl_lower')
        formats(filename, 'teplozap_lower')

    dbConnection = sqlite3.connect('database.db', check_same_thread=True)

    dataFrame_teploz = pds.read_sql(
        "select description as description_teploz,url as url_teploz, id as id_teploz,sku as sku_teploz,\
        article as article_teploz,name as name_teploz,price as price_teploz from "
        "\"teplozapchast_catalog\"",
        dbConnection)
    dataFrame_kotel = pds.read_sql(
        "select description as description_kotel_kr,url as url_kotel, id as id_kotel_kr,sku as sku_kotel_kr,\
        article as article_kotel_kr,name as name_kotel_kr,price as price_kotel_kr from "
        "\"kotel_kr_catalog\"",
        dbConnection)
    dataFrame_piramida = pds.read_sql(
        "select description as description_piramida,url as url_piramida, name,id as id_piramida,sku as sku_piramida,\
        article as article_piramida,name as name_piramida,"
        "price as price_piramida from \"piramida_catalog\"",
        dbConnection)

    dataFrame_gazkomplekt = pds.read_sql(
        "select description as description_gazc,url as url_gazkomplekt, name,id as id_gazkomplekt,sku as sku_gazkomplekt,\
        article as article_gazkomplekt,"
        "name as name_gazkomplekt, "
        "price as price_gazkomplekt from \"gazkomplekt_catalog\"",
        dbConnection)

    # Разделение товаров на отдельные строки с отдельным артикулом, если у товара несколько артикулов и они перечислены
    # через запятую
    dataFrame_teploz = dataFrame_teploz["sku_teploz"].str.split(",").explode().str.strip().to_frame().combine_first(
        dataFrame_teploz)
    dataFrame_article = pds.concat([dataFrame_kotel['sku_kotel_kr'], dataFrame_teploz['sku_teploz']],
                                   ignore_index=False).drop_duplicates()

    list_error_name = ['G1/2', 'G3/4', '- 100', '- 250', 'G3/8', 'G1/2']
    dataFrame_article = dataFrame_article.loc[~dataFrame_article.isin(list_error_name)]

    # Поиск артикулов в фрейме Пирамиды по полю Описание и Название и запись их в отдельную колонку sku
    modify_piramida24_table('description_piramida')
    modify_piramida24_table('name')
    modify_gazkomplekt_table('name')

    # Разделение товаров на отдельные строки с отдельным артикулом, если у товара несколько артикулов и они перечислены
    # через запятую
    dataFrame_teploz = dataFrame_teploz["sku_teploz"].str.split(",").explode().str.strip().to_frame().combine_first(
        dataFrame_teploz)
    dataFrame_piramida = dataFrame_piramida["sku_piramida"].str.split(
        ",").explode().str.strip().to_frame().combine_first(
        dataFrame_piramida)
    dataFrame_gazkomplekt = dataFrame_gazkomplekt["sku_gazkomplekt"].str.split(
        ",").explode().str.strip().to_frame().combine_first(
        dataFrame_gazkomplekt)

    dbConnection.close()

    # Create Excel file
    kotel_kr_teplozap()
    teplozap_piramida()
    teplozap_gazkompl()
